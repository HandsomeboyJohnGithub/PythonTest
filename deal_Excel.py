import openpyxl

##创建一个工作簿对象
wb = openpyxl.load_workbook('./data/test.xlsx')

#获取工作簿的sheet表的名称
print(wb.sheetnames)

#获取指定的sheet对象
sheet =wb['进阶信息']
print(sheet)


##说明： 有了Worksheet对象后，就可以按名字访问Cell对象
#属性：
'''
 value : Cell中存储的值
 row: 行索引
 columen:列索引
 coordinate: 坐标
'''
Cell = sheet['A1']
print(Cell.value)
print(Cell.row)
print(Cell.column)
print(Cell.coordinate)

#####从工作表中取得行和列
##将Worksheet对象进行切片操作，
##从而取得电子表格中一行、一列或一个矩形中的所有Cell对象
##for cell_row in sheet['A2':'D8']:
##    for cell in cell_row:
##        print(cell.coordinate,cell.value)
##        

## 获取到表中的列并取得每一个单元格的值（下标从0开始）:
for  cell in list(sheet.columns)[0]:
    print(cell.value)

for  cell in list(sheet.columns)[1]:
    print(cell.value)
    

for  cell in list(sheet.columns)[2]:
    print(cell.value)


for  cell in list(sheet.columns)[3]:
    print(cell.value)


##获取工作表中的最大行和最大列的数量
print(sheet.max_row,sheet.max_column)
    
